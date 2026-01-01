import tkinter as tk
from tkinter import ttk, scrolledtext, messagebox, filedialog
import threading
import sys
import pandas as pd
from datetime import date, timedelta, datetime
import traceback
import logging

import strategy_backtester
import strategies
import csv

# 嘗試匯入 reader (Data Manager)
try:
    import reader
except ImportError:
    reader = None

# 嘗試匯入 plotter (Chart)
try:
    import plotter
except ImportError:
    plotter = None

# 設定 Logging
def setup_logging():
    # 建立 Logger
    logger = logging.getLogger()
    logger.setLevel(logging.INFO)
    
    # Formatter
    formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
    
    # 1. File Handler (寫入 application.log)
    file_handler = logging.FileHandler("application.log", encoding='utf-8')
    file_handler.setFormatter(formatter)
    logger.addHandler(file_handler)
    
    # 2. Stream Handler (寫入 stdout -> 會被重導向到 GUI Log Area)
    # 為了避免與下面的 stdout 重導向衝突無限迴圈，我們這裡不直接加 StreamHandler 到 root
    # 或者我們只依賴 print() 會被 RedirectText 捕捉
    # 但 logging 預設 output 是 stderr
    
    # 這裡我們手動讓 logging.info 也 print 出來，這樣就會進 GUI
    stream_handler = logging.StreamHandler(sys.stdout)
    stream_handler.setFormatter(formatter)
    logger.addHandler(stream_handler)
    
    logging.info("Logging system initialized.")

# ---------------------------------------------------------
# Util Class for logging redirection
# ---------------------------------------------------------
class RedirectText:
    """將 stdout/stderr 重導向至 Tkinter Text 元件"""
    def __init__(self, text_widget):
        self.text_widget = text_widget

    def write(self, string):
        # 使用 after 確保在主執行緒更新 UI
        if self.text_widget:
            self.text_widget.after(0, lambda: self._append(string))

    def _append(self, string):
        try:
            self.text_widget.configure(state='normal')
            self.text_widget.insert(tk.END, string)
            self.text_widget.see(tk.END)
            self.text_widget.configure(state='disabled')
        except Exception:
            pass

    def flush(self):
        pass

# ---------------------------------------------------------
# Tab 1: Data Manager Frame
# ---------------------------------------------------------
class DataManagerFrame(ttk.Frame):
    def __init__(self, parent):
        super().__init__(parent, padding="10")
        
        if reader is None:
            ttk.Label(self, text="錯誤: 找不到 reader.py", foreground="red").pack()
            return

        # 上方控制區
        control_frame = ttk.Frame(self)
        control_frame.pack(fill=tk.X, pady=5)
        
        ttk.Label(control_frame, text="抓取最近").pack(side=tk.LEFT)
        
        self.days_var = tk.StringVar(value="5")
        self.days_entry = ttk.Entry(control_frame, textvariable=self.days_var, width=5, justify="center")
        self.days_entry.pack(side=tk.LEFT, padx=5)
        
        ttk.Label(control_frame, text="個交易日").pack(side=tk.LEFT)
        
        self.start_btn = ttk.Button(control_frame, text="開始抓取資料", command=self.on_start)
        self.start_btn.pack(side=tk.LEFT, padx=20)
        
        # 狀態顯示
        self.status_label = ttk.Label(control_frame, text="準備就緒", foreground="gray")
        self.status_label.pack(side=tk.LEFT, padx=5)

        # 輸出日誌區
        log_frame = ttk.LabelFrame(self, text=" 執行紀錄 ", padding="5")
        log_frame.pack(fill=tk.BOTH, expand=True, pady=10)
        
        self.log_area = scrolledtext.ScrolledText(log_frame, state='disabled', font=("Consolas", 9))
        self.log_area.pack(fill=tk.BOTH, expand=True)

        # 保存原始 stdout (切換 tab 時可能需要恢復，或由 MainApp 統一管理)
        # 這裡簡單做：當點擊日時暫時綁定
        self.redirector = RedirectText(self.log_area)

    def on_start(self):
        days_str = self.days_var.get()
        if not days_str.isdigit() or int(days_str) <= 0:
            messagebox.showwarning("輸入錯誤", "請輸入大於 0 的整數 days")
            return

        days = int(days_str)
        self.toggle_ui(running=True)
        self.status_label.config(text=f"正在執行 (最近 {days} 天)...", foreground="blue")
        
        # 清空 log
        self.log_area.configure(state='normal')
        self.log_area.delete(1.0, tk.END)
        self.log_area.configure(state='disabled')

        # 綁定 stdout
        sys.stdout = self.redirector
        sys.stderr = self.redirector

        # 啟動執行緒
        threading.Thread(target=self.run_task, args=(days,), daemon=True).start()

    def run_task(self, days):
        try:
            # 建構參數
            class Args: pass
            args = Args()
            args.days = days
            args.date_from = None
            args.date_to = None
            args.sleep = 0.2
            args.max_retries = 3
            args.batch_size = 5000
            args.force = False
            args.log_level = "INFO"
            
            # 使用 reader 模組內的路徑
            args.data_dir = str(reader.DEFAULT_DATA_DIR)
            args.db_path = str(reader.DEFAULT_DB_PATH)
            
            args.out_format = "csv"
            args.from_cache_only = False
            args.refresh_calendar = False
            args.halt_on_fail = 20

            reader.run(args)
            
            self.after(0, lambda: self.finish_task(success=True))
            
        except Exception as e:
            logging.error("Data Fetch Task Failed", exc_info=True)
            err_msg = str(e)
            self.after(0, lambda: self.finish_task(success=False, error_msg=err_msg))

    def finish_task(self, success, error_msg=None):
        self.toggle_ui(running=False)
        # 恢復 stdout (可選，視需求)
        # sys.stdout = sys.__stdout__
        
        if success:
            self.status_label.config(text="執行完成", foreground="green")
            messagebox.showinfo("完成", "資料抓取與入庫完成！")
        else:
            self.status_label.config(text="執行失敗", foreground="red")
            messagebox.showerror("錯誤", f"執行過程發生錯誤：\n{error_msg}\n\n詳細錯誤已寫入 errorlog.txt")

    def toggle_ui(self, running):
        state = 'disabled' if running else 'normal'
        self.start_btn.config(state=state)
        self.days_entry.config(state=state)


# ---------------------------------------------------------
# Tab 2: Strategy Backtester Frame
# ---------------------------------------------------------
class StrategyFrame(ttk.Frame):
    def __init__(self, parent):
        super().__init__(parent, padding="10")
        
        # --- Top Control Panel ---
        control_frame = ttk.LabelFrame(self, text="策略設定", padding="10")
        control_frame.pack(fill=tk.X, pady=5)
        
        # Strategy Helper
        lbl_select = ttk.Label(control_frame, text="選擇策略:")
        lbl_select.grid(row=0, column=0, padx=5, pady=2, sticky="w")
        
        self.btn_toggle_select = ttk.Button(control_frame, text="全選/取消", width=10, command=self.toggle_select_all)
        self.btn_toggle_select.grid(row=1, column=0, padx=5, pady=2, sticky="nw")
        
        self.strategy_listbox = tk.Listbox(control_frame, selectmode=tk.MULTIPLE, height=4, width=40, exportselection=0)
        strategies_list = strategies.DAILY_STRATEGIES
        for s in strategies_list:
            self.strategy_listbox.insert(tk.END, s)
        self.strategy_listbox.select_set(0) # 預設選中第一個
        
        vsb = ttk.Scrollbar(control_frame, orient=tk.VERTICAL, command=self.strategy_listbox.yview)
        hsb = ttk.Scrollbar(control_frame, orient=tk.HORIZONTAL, command=self.strategy_listbox.xview)
        self.strategy_listbox.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        
        self.strategy_listbox.grid(row=0, column=1, rowspan=2, padx=5, pady=5, sticky="nsew")
        vsb.grid(row=0, column=2, rowspan=2, sticky="ns", pady=5)
        hsb.grid(row=2, column=1, sticky="ew", padx=5)
        
        # Mode Selection
        ttk.Label(control_frame, text="執行模式:").grid(row=0, column=3, padx=10, pady=5, sticky="w")
        self.backtest_var = tk.BooleanVar(value=False)
        self.chk_backtest = ttk.Checkbutton(control_frame, text="回測", variable=self.backtest_var, command=self.toggle_dates)
        self.chk_backtest.grid(row=0, column=4, padx=5, pady=5, sticky="w")
        
        # Date Range
        self.date_frame = ttk.Frame(control_frame)
        self.date_frame.grid(row=1, column=3, columnspan=2, sticky="w", padx=5, pady=5)
        
        ttk.Label(self.date_frame, text="範圍:").pack(side=tk.LEFT)
        self.start_entry = ttk.Entry(self.date_frame, width=11)
        self.start_entry.insert(0, (date.today() - timedelta(days=90)).strftime('%Y-%m-%d'))
        self.start_entry.pack(side=tk.LEFT, padx=2)
        
        ttk.Label(self.date_frame, text="~").pack(side=tk.LEFT)
        self.end_entry = ttk.Entry(self.date_frame, width=11)
        self.end_entry.insert(0, date.today().strftime('%Y-%m-%d'))
        self.end_entry.pack(side=tk.LEFT, padx=2)
        
        # Buttons Frame
        btn_frame = ttk.Frame(control_frame)
        btn_frame.grid(row=0, column=5, rowspan=2, padx=20, sticky="nesw")
        
        self.btn_run = ttk.Button(btn_frame, text="執行策略", command=self.on_run)
        self.btn_run.pack(fill=tk.X, pady=2)
        
        self.btn_plot = ttk.Button(btn_frame, text="畫 K 線圖", command=self.on_plot)
        self.btn_plot.pack(fill=tk.X, pady=2)
        
        self.btn_export = ttk.Button(btn_frame, text="匯出結果", command=self.on_export)
        self.btn_export.pack(fill=tk.X, pady=2)
        
        # 暫存結果與 Meta 資訊
        self.current_results = pd.DataFrame()
        self.current_meta = {
            'start_date': '',
            'end_date': '',
            'strategies': []
        }
        
        # Status
        self.status_var = tk.StringVar(value="就緒")
        self.lbl_status = ttk.Label(control_frame, textvariable=self.status_var, foreground="blue")
        self.lbl_status.grid(row=2, column=0, columnspan=3, sticky="w", padx=5)

        # Progress Bar
        self.progress = ttk.Progressbar(control_frame, length=200, mode='determinate')
        self.progress.grid(row=2, column=3, columnspan=2, sticky="ew", padx=5)

        # --- Results Area ---
        result_frame = ttk.Frame(self, padding="5")
        result_frame.pack(fill=tk.BOTH, expand=True)
        
        cols = ('策略', '代號', '名稱', '訊號日期', '收盤價', '買入日期', '買入價', '報酬5日', '報酬10日', '報酬20日', '報酬60日')
        self.tree = ttk.Treeview(result_frame, columns=cols, show='headings')
        
        # Define headings
        for col in cols:
            self.tree.heading(col, text=col, command=lambda c=col: self.treeview_sort_column(c, False))
            self.tree.column(col, width=90, anchor="center")
        self.tree.column('名稱', width=120)
        
        # Scrollbars
        vsb = ttk.Scrollbar(result_frame, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(result_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        
        self.tree.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')
        
        result_frame.grid_columnconfigure(0, weight=1)
        result_frame.grid_rowconfigure(0, weight=1)
        
        # --- Performance Summary Area ---
        self.summary_frame = ttk.LabelFrame(self, text=" 績效摘要 (Performance Summary) ", padding="10")
        self.summary_frame.pack(fill=tk.X, pady=5)
        
        # Labels for summary (organized in a grid)
        self.summary_labels = {}
        # Signal Count
        ttk.Label(self.summary_frame, text="總訊號數:").grid(row=0, column=0, padx=5, sticky="w")
        self.summary_labels['count'] = ttk.Label(self.summary_frame, text="0", foreground="blue", font=("Arial", 10, "bold"))
        self.summary_labels['count'].grid(row=0, column=1, padx=20, sticky="w")
        
        # Table-like header for returns
        headers = ["週期", "平均報酬", "勝率", "獲利因子", "期望值", "最大回撤", "連虧"]
        for i, h in enumerate(headers):
            ttk.Label(self.summary_frame, text=h, font=("Arial", 9, "bold")).grid(row=0, column=i+2, padx=10)
            
        self.stats_rows = [] # To store the dynamically updated labels
        # We will initialize placeholders for 4 return periods
        for i in range(4):
            period_lbl = ttk.Label(self.summary_frame, text="-")
            avg_lbl = ttk.Label(self.summary_frame, text="-", foreground="darkred")
            win_lbl = ttk.Label(self.summary_frame, text="-", foreground="darkgreen")
            pf_lbl = ttk.Label(self.summary_frame, text="-", foreground="blue")
            exp_lbl = ttk.Label(self.summary_frame, text="-", foreground="purple")
            mdd_lbl = ttk.Label(self.summary_frame, text="-", foreground="red")
            con_lbl = ttk.Label(self.summary_frame, text="-", foreground="orange")
            
            period_lbl.grid(row=i+1, column=2, padx=10)
            avg_lbl.grid(row=i+1, column=3, padx=10)
            win_lbl.grid(row=i+1, column=4, padx=10)
            pf_lbl.grid(row=i+1, column=5, padx=10)
            exp_lbl.grid(row=i+1, column=6, padx=10)
            mdd_lbl.grid(row=i+1, column=7, padx=10)
            con_lbl.grid(row=i+1, column=8, padx=10)
            self.stats_rows.append((period_lbl, avg_lbl, win_lbl, pf_lbl, exp_lbl, mdd_lbl, con_lbl))

        # Initialize Logic
        self.backtester = strategy_backtester.StrategyBacktester()
        if plotter:
            self.plotter = plotter.StockPlotter()
        else:
            self.plotter = None
            
        self.toggle_dates()
    
    def treeview_sort_column(self, col, reverse):
        """點擊標題排序功能"""
        l = [(self.tree.set(k, col), k) for k in self.tree.get_children('')]
        def convert(val):
            if val in ('-', 'N/A', '', 'None'): return float('-inf')
            try: return float(val)
            except ValueError: return str(val)
        try:
            l.sort(key=lambda t: convert(t[0]), reverse=reverse)
        except TypeError:
            l.sort(key=lambda t: str(t[0]), reverse=reverse)
        for index, (val, k) in enumerate(l):
            self.tree.move(k, '', index)
        self.tree.heading(col, command=lambda: self.treeview_sort_column(col, not reverse))

    def toggle_select_all(self):
        """切換全選或全取消選取"""
        current_selection = self.strategy_listbox.curselection()
        total_items = self.strategy_listbox.size()
        
        if len(current_selection) == total_items:
            # 如果已經全選，則全部取消
            self.strategy_listbox.selection_clear(0, tk.END)
        else:
            # 否則全部選取
            self.strategy_listbox.selection_set(0, tk.END)

    def toggle_dates(self):
        if self.backtest_var.get():
            self.start_entry.config(state='normal')
            self.end_entry.config(state='normal')
        else:
            self.start_entry.config(state='disabled')
            self.end_entry.config(state='disabled')

    def update_progress(self, current, total):
        """Update progress bar in main thread"""
        if total > 0:
            pct = (current / total) * 100
            self.progress['value'] = pct
            self.root.update_idletasks() # Force update

    def on_run(self):
        # 取得所有選中的策略
        selected_indices = self.strategy_listbox.curselection()
        if not selected_indices:
            messagebox.showwarning("提示", "請至少選擇一個策略")
            return
        
        strategies = [self.strategy_listbox.get(i) for i in selected_indices]
        latest_only = not self.backtest_var.get()
        start = self.start_entry.get()
        end = self.end_entry.get()
        
        # 儲存 Meta 資訊供匯出使用
        self.current_meta = {
            'start_date': '-' if latest_only else start,
            'end_date': '-' if latest_only else end,
            'strategies': strategies
        }
        
        if not latest_only:
            try:
                pd.to_datetime(start)
                pd.to_datetime(end)
            except:
                messagebox.showerror("錯誤", "日期格式不正確，請使用 YYYY-MM-DD")
                return

        self.btn_run.config(state='disabled')
        self.tree.delete(*self.tree.get_children())
        self.status_var.set("正在掃描資料庫，請稍候...")
        self.progress['value'] = 0
        logging.info(f"Starting strategy run: {strategies}, LatestOnly={latest_only}, Range={start}-{end}")
        threading.Thread(target=self.run_task, args=(strategies, latest_only, start, end), daemon=True).start()

    def run_task(self, strategies, latest_only, start, end):
        try:
            # Callback handler for passing to backtester
            def progress_handler(current, total):
                self.after(0, lambda: self.update_progress(current, total))
                
            df_res = self.backtester.run_scan(strategies, latest_only, start, end, progress_callback=progress_handler)
            logging.info(f"Scan complete. Signals found: {len(df_res)}")
            self.after(0, lambda: self.show_results(df_res))
        except Exception as e:
            logging.error("Strategy execution failed", exc_info=True)
            err_msg = str(e)
            self.after(0, lambda: self.show_error(err_msg))

    def show_results(self, df):
        self.status_var.set(f"掃描完成，共找到 {len(df)} 筆訊號")
        self.progress['value'] = 100
        self.btn_run.config(state='normal')
        
        # 儲存結果
        self.current_results = df.copy()
        
        if df.empty:
            # 清空摘要
            self.update_summary(pd.DataFrame())
            messagebox.showinfo("結果", "在此條件下未發現任何訊號。")
            return
            
        df = df.fillna("-")
        for _, row in df.iterrows():
            vals = [row[col] for col in self.tree['columns']]
            self.tree.insert("", "end", values=vals)
        
        # 更新績效摘要
        self.update_summary(df)

    def update_summary(self, df):
        """計算並顯示績效摘要"""
        if df.empty:
            self.summary_labels['count'].config(text="0")
            for p_lbl, a_lbl, w_lbl, pf_lbl, exp_lbl, mdd_lbl, con_lbl in self.stats_rows:
                p_lbl.config(text="-")
                a_lbl.config(text="-")
                w_lbl.config(text="-")
                pf_lbl.config(text="-")
                exp_lbl.config(text="-")
                mdd_lbl.config(text="-")
                con_lbl.config(text="-")
            return
            
        # 1. 總訊號數
        self.summary_labels['count'].config(text=str(len(df)))
        
        # 2. 找出報酬欄位 (過濾掉 '策略', '代號' 等)
        # 日線通常是 '報酬5日', '報酬10日'... 週線是 '報酬5週'... 
        return_cols = [c for c in df.columns if '報酬' in c]
        
        for i, col in enumerate(return_cols):
            if i >= len(self.stats_rows): break
            
            p_lbl, a_lbl, w_lbl, pf_lbl, exp_lbl, mdd_lbl, con_lbl = self.stats_rows[i]
            
            # 過濾無效數據
            valid_series = df[col].apply(lambda x: pd.to_numeric(x, errors='coerce'))
            valid_data = valid_series.dropna()
            
            if not valid_data.empty:
                avg_ret = valid_data.mean()
                win_data = valid_data[valid_data > 0]
                loss_data = valid_data[valid_data <= 0]
                win_rate = len(win_data) / len(valid_data) * 100
                
                # 1. 獲利因子 (Profit Factor) = 總獲利 / 總虧損(絕對值)
                total_profit = win_data.sum()
                total_loss = abs(loss_data.sum())
                profit_factor = total_profit / total_loss if total_loss > 0 else (float('inf') if total_profit > 0 else 0)
                
                # 2. 期望值 (Expectancy) = (勝率 * 平均獲利) - (敗率 * 平均虧損)
                avg_profit = win_data.mean() if not win_data.empty else 0
                avg_loss = abs(loss_data.mean()) if not loss_data.empty else 0
                loss_rate = len(loss_data) / len(valid_data)
                win_rate_dec = len(win_data) / len(valid_data)
                expectancy = (win_rate_dec * avg_profit) - (loss_rate * avg_loss)
                
                # 3. 最大回撤 (MDD) - 依據訊號日期排序後計算累計報酬
                sorted_df = df.iloc[valid_data.index].copy()
                sorted_df['訊號日期_dt'] = pd.to_datetime(sorted_df['訊號日期'])
                sorted_df = sorted_df.sort_values('訊號日期_dt')
                returns_seq = sorted_df[col].apply(lambda x: pd.to_numeric(x, errors='coerce')).dropna()
                # 改用「單利累加」避免訊號過多時的乘數效應
                equity = 100 + returns_seq.cumsum()
                peak = equity.cummax()
                drawdown = (equity - peak) / peak
                mdd = drawdown.min() * 100 if not drawdown.empty else 0
                
                # 4. 最大連續虧損 (Max Consecutive Losses)
                is_loss = returns_seq <= 0
                consecutive_losses = is_loss.astype(int).groupby((is_loss != is_loss.shift()).cumsum()).cumsum().max()
                
                # 更新 Label
                p_lbl.config(text=col.replace("報酬", ""))
                a_lbl.config(text=f"{avg_ret:+.2f}%", foreground="darkred" if avg_ret > 0 else "darkgreen")
                w_lbl.config(text=f"{win_rate:.1f}%")
                
                pf_text = f"{profit_factor:.2f}" if profit_factor != float('inf') else "∞"
                pf_lbl.config(text=pf_text, foreground="blue" if profit_factor >= 1.5 else "black")
                
                exp_lbl.config(text=f"{expectancy:+.2f}%", foreground="purple" if expectancy > 0 else "black")
                
                mdd_lbl.config(text=f"{mdd:+.2f}%", foreground="red" if mdd < -15 else "black")
                con_lbl.config(text=str(int(consecutive_losses)) if not pd.isna(consecutive_losses) else "0")
            else:
                p_lbl.config(text=col.replace("報酬", ""))
                a_lbl.config(text="N/A", foreground="gray")
                w_lbl.config(text="0.0%")
                pf_lbl.config(text="N/A", foreground="black")
                exp_lbl.config(text="N/A", foreground="black")
                mdd_lbl.config(text="N/A", foreground="black")
                con_lbl.config(text="N/A", foreground="black")

    def show_error(self, msg):
        self.status_var.set("發生錯誤")
        self.btn_run.config(state='normal')
        messagebox.showerror("執行錯誤", f"{msg}\n\n詳細錯誤已寫入 errorlog.txt")

    def on_export(self):
        """匯出回測明細與績效摘要"""
        if self.current_results.empty:
            messagebox.showwarning("提示", "目前沒有回測數據可供匯出")
            return
            
        # 1. 準備預設檔名: [第一個策略]_[日期時間]
        first_strat = self.current_meta['strategies'][0] if self.current_meta['strategies'] else "Strategy"
        now_str = datetime.now().strftime("%Y%m%d_%H%M")
        default_filename = f"{first_strat}_{now_str}"
        
        # 2. 選擇路徑
        file_path = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV Files", "*.csv"), ("Excel Files", "*.xlsx")],
            initialfile=default_filename,
            title="選擇儲存路徑"
        )
        if not file_path:
            return
            
        try:
            # 3. 準備 Meta 與摘要內容
            meta_header = [
                ["--- 回測參數 (Backtest Parameters) ---"],
                ["回測時間範圍", f"{self.current_meta['start_date']} ~ {self.current_meta['end_date']}"],
                ["選用策略清單", ", ".join(self.current_meta['strategies'])],
                ["匯出時間", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
                [] # 空行
            ]
            
            summary_header = [
                ["--- 績效摘要 (Performance Summary) ---"],
                ["總訊號數", self.summary_labels['count'].cget("text")]
            ]
            
            # 擷取表格狀摘要
            summary_table_headers = ["週期", "平均報酬", "勝率", "獲利因子", "期望值", "最大回撤", "連虧"]
            summary_header.append(summary_table_headers)
            for row in self.stats_rows:
                period = row[0].cget("text")
                if period == "-" or period == "": continue
                summary_header.append([lbl.cget("text") for lbl in row])
            
            summary_header.append([]) # 空行
            summary_header.append(["--- 交易明細 (Trade Details) ---"])
            
            # 4. 執行寫入
            if file_path.lower().endswith('.csv'):
                with open(file_path, 'w', newline='', encoding='utf-8-sig') as f:
                    writer = csv.writer(f)
                    writer.writerows(meta_header)
                    writer.writerows(summary_header)
                    # 寫入明細標題
                    writer.writerow(self.current_results.columns.tolist())
                    # 寫入明細資料
                    writer.writerows(self.current_results.values.tolist())
            else:
                # Excel 匯出
                try:
                    import openpyxl
                    from openpyxl.utils.dataframe import dataframe_to_rows
                    
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = "回測報告"
                    
                    # 寫入 Meta & 摘要
                    for r_idx, r_data in enumerate(meta_header + summary_header, 1):
                        for c_idx, value in enumerate(r_data, 1):
                            ws.cell(row=r_idx, column=c_idx, value=value)
                    
                    # 寫入明細 (接在摘要後面)
                    start_row = len(meta_header) + len(summary_header) + 1
                    # 標題
                    for c_idx, col in enumerate(self.current_results.columns, 1):
                        ws.cell(row=start_row, column=c_idx, value=col)
                    # 內容
                    for r_idx, row_data in enumerate(self.current_results.values, 1):
                        for c_idx, value in enumerate(row_data, 1):
                            ws.cell(row=start_row + r_idx, column=c_idx, value=value)
                            
                    wb.save(file_path)
                except ImportError:
                    # 如果沒裝 openpyxl，降級使用 pandas 內建但不一定成功的 excel 寫入或提示
                    messagebox.showwarning("提示", "未偵測到 openpyxl 模組，改以 CSV 格式寫入 Meta，明細則儲存在另一頁。")
                    with pd.ExcelWriter(file_path) as writer:
                        self.current_results.to_excel(writer, index=False, sheet_name='交易明細')
                
            messagebox.showinfo("成功", f"回測結果已成功儲存至：\n{file_path}")
            
        except Exception as e:
            logging.error(f"Export failed: {e}", exc_info=True)
            messagebox.showerror("匯出失敗", f"寫入檔案時發生錯誤：\n{e}")

    def on_plot(self):
        selected_item = self.tree.selection()
        if not selected_item:
            messagebox.showwarning("提示", "請先從列表中選擇一檔股票")
            return
        item = self.tree.item(selected_item[0])
        values = item['values']
        if not values: return
        code = str(values[1])
        name = str(values[2])
        signal_date = str(values[3])
        
        if self.plotter is None:
            messagebox.showerror("錯誤", "尚未安裝 matplotlib，無法使用畫圖功能。\n請執行 install_requirements.bat 進行安裝。")
            return
        try:
            self.plotter.show_chart(self.winfo_toplevel(), code, name, signal_date=signal_date)
        except Exception as e:
            logging.error("Plotting failed", exc_info=True)
            messagebox.showerror("繪圖錯誤", f"{str(e)}\n\n詳細錯誤已寫入 application.log")


# ---------------------------------------------------------
# Tab 3: Weekly Strategy Backtester Frame
# ---------------------------------------------------------
class WeeklyStrategyFrame(StrategyFrame):
    """
    繼承 StrategyFrame 但覆寫執行邏輯以支援週線策略
    """
    def __init__(self, parent):
        super().__init__(parent)
        
        # 修改按鈕跟 Label
        # 由於繼承自 StrategyFrame，已經有 self.strategy_cb, self.tree 等元件
        # 1. 更新策略列表
        self.strategy_listbox.delete(0, tk.END)
        strategies_list = strategies.WEEKLY_STRATEGIES
        for s in strategies_list:
            self.strategy_listbox.insert(tk.END, s)
        self.strategy_listbox.select_set(0)
        
        # 2. 更新 Treeview 欄位 (週)
        cols = ('策略', '代號', '名稱', '訊號日期', '收盤價', '買入日期(週)', '買入價', '報酬5週', '報酬10週', '報酬20週', '報酬60週')
        self.tree['columns'] = cols
        for col in cols:
            self.tree.heading(col, text=col, command=lambda c=col: self.treeview_sort_column(c, False))
            self.tree.column(col, width=90, anchor="center")
        self.tree.column('策略', width=180) # 策略名稱通常較長
        self.tree.column('名稱', width=100)
        
        # 禁用 "畫圖" 或 "最新訊號" 若不適用
        # 週線目前還是可以用 "最新訊號"，只看最後一週
        # 畫圖功能若只支援日線，可能需要改寫 plotter 支援週線或禁用
        # 這裡暫時保留按鈕，但提示使用者畫圖仍是日線
        
    def run_task(self, strategies, latest_only, start, end):
        """
        覆寫執行緒任務: 呼叫 run_weekly_scan
        """
        try:
            # 這裡不支援 latest_only (或需要 backtester 支援)
            # 這裡簡單處理：若 latest_only=True，可以透過 run_weekly_scan 後只取最後
            
            # 但 backtester.run_weekly_scan 目前沒實作 latest_only 參數
            # 我們可以都在此做 filtering
            
            # Callback handler
            def progress_handler(current, total):
                self.after(0, lambda: self.update_progress(current, total))

            logging.info(f"Running weekly strategies: {strategies}, Range={start}-{end}")
            df_res = self.backtester.run_weekly_scan(strategies, start, end, progress_callback=progress_handler)
            logging.info(f"Weekly scan complete. Signals: {len(df_res)}")
            
            if latest_only and not df_res.empty:
                # [Fix] 實作 latest_only 過濾
                # 找出結果中最新的日期 (通常是最近一個週五)
                max_date = df_res['訊號日期'].max()
                df_res = df_res[df_res['訊號日期'] == max_date]

            self.after(0, lambda: self.show_results(df_res))
        except Exception as e:
            logging.error("Weekly strategy failed", exc_info=True)
            err_msg = str(e)
            self.after(0, lambda: self.show_error(err_msg))
            
    def on_plot(self):
        selected_item = self.tree.selection()
        if not selected_item:
            messagebox.showwarning("提示", "請先從列表中選擇一檔股票")
            return
            
        item = self.tree.item(selected_item[0])
        values = item['values']
        if not values: return
            
        code = str(values[1])
        name = str(values[2])
        signal_date = str(values[3]) # 這是週五日期
        
        if self.plotter is None:
            messagebox.showerror("錯誤", "尚未安裝 matplotlib")
            return
            
        try:
            # 呼叫 Plotter 顯示週線圖
            self.plotter.show_chart(self.winfo_toplevel(), code, name, signal_date=signal_date, frequency='W')
        except Exception as e:
            logging.error("Plotting failed", exc_info=True)
            messagebox.showerror("繪圖錯誤", f"{str(e)}\n\n詳細錯誤已寫入 application.log")


# ---------------------------------------------------------
# Main App (Tabbed Interface)
# ---------------------------------------------------------
class TWSEApp:
    def __init__(self, root):
        self.root = root
        self.root.title("TWSE 資料管理與策略回測系統")
        self.root.geometry("1150x700")

        # 樣式調整
        style = ttk.Style()
        try:
            style.theme_use('vista')
        except:
            pass
        style.configure("TButton", font=("Microsoft JhengHei", 10))
        style.configure("TLabel", font=("Microsoft JhengHei", 11))

        # 建立 Notebook (分頁)
        self.notebook = ttk.Notebook(root)
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # Tab 1: 資料管理
        self.tab_reader = DataManagerFrame(self.notebook)
        self.notebook.add(self.tab_reader, text="資料管理 (Reader)")

        # Tab 2: 策略回測 (日)
        self.tab_strategy = StrategyFrame(self.notebook)
        self.notebook.add(self.tab_strategy, text="策略回測 (日K)")

        # Tab 3: 策略回測 (週)
        self.tab_weekly = WeeklyStrategyFrame(self.notebook)
        self.notebook.add(self.tab_weekly, text="策略回測 (週K)")


if __name__ == "__main__":
    setup_logging()
    root = tk.Tk()
    app = TWSEApp(root)
    root.mainloop()
